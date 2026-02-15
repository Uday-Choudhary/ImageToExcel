"""
Excel Converter
Converts OCR JSON output to Excel using SpatialTableExtractor.
"""

import os
import re
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from spatial_table_extractor import SpatialTableExtractor

OCR_DATA_DIR = "ocr_data"
OUTPUT_DIR = "output"

# Validation colors
COLOR_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
COLOR_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def save_to_excel(base_dir=".", output_file="Extracted_Data_OCR.xlsx"):
    """
    Process all EasyOCR JSON files from ocr_data/ and save
    extracted header info, tables, and footer info to output/Extracted_Data.xlsx.
    """
    data_dir = os.path.join(base_dir, OCR_DATA_DIR)
    out_dir = os.path.join(base_dir, OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)

    if not os.path.isdir(data_dir):
        print(f"  No {OCR_DATA_DIR}/ folder found. Run OCR first.")
        return None

    json_files = sorted(
        f for f in os.listdir(data_dir) if f.endswith('_easyocr.json')
    )
    if not json_files:
        print(f"  No JSON files in {OCR_DATA_DIR}/.")
        return None

    extractor = SpatialTableExtractor()
    output_path = os.path.join(out_dir, output_file)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        count = 0
        for fname in json_files:
            json_path = os.path.join(data_dir, fname)
            print(f"  {fname}")
            
            # Use new method to get full data
            data = extractor.extract_full_data(json_path)

            if not data or not data.get("table") or not data["table"].get("rows"):
                print(f"    → No structured data found, skipping")
                continue

            header_split = data.get("header_split", {"left": [], "right": []})
            # fallback for old extractor version
            if not header_split.get("left") and not header_split.get("right"):
                # try finding old header_info
                flat_header = data.get("header_info", [])
                header_split["left"] = [ " ".join(r) for r in flat_header ]

            metadata = data.get("metadata", {})
            table = data["table"]
            footer_info = data.get("footer_info", [])
            
            headers = table["headers"]
            rows = table["rows"]

            # Sheet name from filename (max 31 chars)
            sheet = fname.replace("_easyocr.json", "")[:31]
            
            # Create DataFrame for the table part
            df = pd.DataFrame(rows, columns=headers)
            
            # Determine start row for table (Header Height + Padding)
            header_height = max(len(header_split["left"]), len(header_split["right"]))
            start_row = max(header_height + 4, 6) # Minimum 6 rows for header space
            
            df.to_excel(writer, sheet_name=sheet, index=False, startrow=start_row)
            
            ws = writer.sheets[sheet]
            
            # --- Write Split Header ---
            # Left Header (Col A)
            for idx, text in enumerate(header_split["left"]):
                cell = ws.cell(row=idx + 1, column=1, value=text)
                cell.font = Font(bold=(idx == 0)) # Bold first line

            # Right Header (Col F or E, depending on table width, default F=6)
            right_col_idx = max(len(headers), 6) 
            for idx, text in enumerate(header_split["right"]):
                cell = ws.cell(row=idx + 1, column=right_col_idx, value=text)
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(bold=(idx == 0))

            # --- Write Metadata Box ---
            # Place metadata in a box to the right of Left Header, or somewhere identifying
            # Let's put it in Column 3 (C) row 1 if space permits, or integrated
            if metadata:
                meta_row = 1
                meta_col = 3 # Col C
                ws.cell(row=meta_row, column=meta_col, value="METADATA EXTRACTED").font = Font(bold=True, underline="single")
                for k, v in metadata.items():
                    meta_row += 1
                    ws.cell(row=meta_row, column=meta_col, value=f"{k.replace('_', ' ').title()}:")
                    ws.cell(row=meta_row, column=meta_col+1, value=v).font = Font(bold=True)
            
            # --- Style Table Headers ---
            header_row_idx = start_row + 1
            for col_idx, col_name in enumerate(headers):
                cell = ws.cell(row=header_row_idx, column=col_idx+1)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # --- Math Validation ---
            # Identify Qty, Price, Total columns
            col_map = {name.lower(): i for i, name in enumerate(headers)}
            
            qty_idx = -1
            price_idx = -1
            total_idx = -1

            for name, idx in col_map.items():
                if any(x in name for x in ['qty', 'quantity', 'units']): qty_idx = idx
                if any(x in name for x in ['price', 'rate', 'unit']): price_idx = idx
                if any(x in name for x in ['total', 'amount', 'net']): total_idx = idx

            # Only validate if we have all three
            if qty_idx != -1 and price_idx != -1 and total_idx != -1:
                # Add a "Validation" column header
                val_col_idx = len(headers) + 1
                ws.cell(row=header_row_idx, column=val_col_idx, value="Validation").font = Font(bold=True)
                
                for r_i, row_data in enumerate(rows):
                    current_row = header_row_idx + 1 + r_i
                    try:
                        q_str = str(row_data[qty_idx]); p_str = str(row_data[price_idx]); t_str = str(row_data[total_idx])
                        q = float(re.sub(r'[^\d.]', '', q_str)) if re.search(r'\d', q_str) else 0
                        p = float(re.sub(r'[^\d.]', '', p_str)) if re.search(r'\d', p_str) else 0
                        t = float(re.sub(r'[^\d.]', '', t_str)) if re.search(r'\d', t_str) else 0
                        
                        # Calculation
                        calc = q * p
                        diff = abs(calc - t)
                        
                        # Threshold $0.05
                        status_cell = ws.cell(row=current_row, column=val_col_idx)
                        if diff < 0.05 and calc > 0:
                            status_cell.value = "OK"
                            status_cell.fill = COLOR_GREEN
                        elif calc > 0:
                            status_cell.value = f"Mismatch (Calc: {calc:.2f})"
                            status_cell.fill = COLOR_RED
                            # Also highlight the Total cell
                            ws.cell(row=current_row, column=total_idx+1).fill = COLOR_RED
                    except:
                        pass

            # --- Write Footer Info ---
            footer_start_row = header_row_idx + len(rows) + 2
            for idx, info_row in enumerate(footer_info):
                target_col_val = len(headers)
                target_col_label = max(1, target_col_val - 1)
                row_num = footer_start_row + idx
                
                if len(info_row) >= 2:
                    ws.cell(row=row_num, column=target_col_label, value=info_row[0]).alignment = Alignment(horizontal='right')
                    ws.cell(row=row_num, column=target_col_val, value=info_row[1]).alignment = Alignment(horizontal='right')
                    # Make Total Due bold
                    if "total" in info_row[0].lower() or "due" in info_row[0].lower():
                        ws.cell(row=row_num, column=target_col_label).font = Font(bold=True)
                        ws.cell(row=row_num, column=target_col_val).font = Font(bold=True)
                else:
                    ws.cell(row=row_num, column=1, value=info_row[0])

            # Auto-size columns
            for col_idx in range(1, len(headers) + 2): # +2 to include Validation col
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' # simplistic
                max_len = 0
                for cell in ws[col_letter]:
                     if cell.value:
                         max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                
            left_count = len(header_split['left']) if header_split.get('left') else 0
            right_count = len(header_split['right']) if header_split.get('right') else 0
            print(f"    → Sheet '{sheet}' | {left_count} Left / {right_count} Right Hdr | {len(rows)} Rows")
            count += 1

    if count > 0:
        print(f"\n  Saved {count} sheet(s) → {output_path}")
        return output_path
    else:
        print("\n  No data extracted.")
        return None


if __name__ == "__main__":
    save_to_excel()
