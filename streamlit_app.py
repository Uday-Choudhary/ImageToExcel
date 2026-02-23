"""
ImageToExcel - Streamlit App
Upload images of tables/invoices and get a structured Excel file back.
Powered by Llama Vision (via Groq) + openpyxl.
"""

import os
import io
import json
import time
import base64
import tempfile

import streamlit as st
import pandas as pd
import numpy as np
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
from groq import Groq

# EasyOCR imports (loaded lazily so app starts fast)
import importlib

# ── Load .env (works locally; on Streamlit Cloud use st.secrets) ──────────────
load_dotenv()

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ImageToExcel — Table Extraction",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');

    /* Base font only — config.toml handles background + text colors */
    html, body, [class*="css"], [class*="st-"] {
        font-family: 'Inter', sans-serif !important;
    }

    /* Headings — firm dark slate for strong hierarchy */
    h1, h2, h3, h4, h5, h6 {
        color: #0f172a !important;
        font-weight: 600 !important;
        letter-spacing: -0.015em;
    }

    /* Body text — muted slate instead of harsh #000 */
    p, span, li,
    div[data-testid="stMarkdownContainer"] p,
    div[data-testid="stMarkdownContainer"] span,
    div[data-testid="stMarkdownContainer"] li {
        color: #475569 !important;
    }

    /* Caption / helper text */
    small, .stCaption p, [data-testid="stCaptionContainer"] p {
        color: #94a3b8 !important;
        font-size: 0.85rem !important;
    }

    /* Layout polish */
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
        max-width: 1200px !important;
    }
    header[data-testid="stHeader"] { display: none !important; }
    footer { display: none !important; }

    /* === HERO === */
    .hero {
        padding: 0 0 2rem;
        border-bottom: 1px solid #e2e8f0;
        margin-bottom: 2.5rem;
    }
    .hero h1 { font-size: 2.25rem !important; color: #0f172a !important; margin-bottom: 0.4rem; }
    .hero p  { font-size: 1.05rem; color: #64748b !important; line-height: 1.65; max-width: 680px; }

    /* === SIDEBAR === */
    [data-testid="stSidebar"] {
        background-color: #ffffff !important;
        border-right: 1px solid #e2e8f0 !important;
    }
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 { color: #0f172a !important; }
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] li { color: #64748b !important; font-size: 0.9rem !important; }

    /* === FILE UPLOADER === */
    [data-testid="stFileUploader"] {
        background-color: #ffffff !important;
        border: 1px dashed #cbd5e1 !important;
        border-radius: 8px !important;
        padding: 2.5rem 1rem !important;
        box-shadow: none !important;
        transition: border-color 0.2s ease;
    }
    [data-testid="stFileUploader"]:hover { border-color: #3b82f6 !important; }
    [data-testid="stFileUploader"] p { color: #94a3b8 !important; }

    /* === BUTTONS (default) === */
    .stButton > button {
        background-color: #ffffff !important;
        color: #0f172a !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 6px !important;
        padding: 0.5rem 1rem !important;
        font-weight: 500 !important;
        font-size: 0.925rem !important;
        transition: all 0.15s ease !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    .stButton > button:hover { border-color: #94a3b8 !important; background-color: #f8fafc !important; }
    .stButton > button:active { transform: translateY(1px); }
    .stButton > button p { color: inherit !important; }

    /* Primary button */
    .stButton > button[kind="primary"] {
        background-color: #2563eb !important;
        color: #ffffff !important;
        border-color: #2563eb !important;
        box-shadow: 0 1px 3px rgba(37,99,235,0.35) !important;
    }
    .stButton > button[kind="primary"] p { color: #ffffff !important; }
    .stButton > button[kind="primary"]:hover {
        background-color: #1d4ed8 !important;
        border-color: #1d4ed8 !important;
    }

    /* Download button */
    .stDownloadButton > button {
        background-color: #f8fafc !important;
        color: #0f172a !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 6px !important;
        font-weight: 500 !important;
        padding: 0.55rem 1rem !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    .stDownloadButton > button p { color: #0f172a !important; }
    .stDownloadButton > button:hover { border-color: #94a3b8 !important; }

    /* === STATUS ALERTS (target Streamlit emotion classes) === */
    div[class*="stAlert"] {
        border-radius: 6px !important;
        border: 1px solid !important;
        padding: 0.75rem 1rem !important;
    }
    div[class*="stSuccess"] { background-color: #f0fdf4 !important; border-color: #bbf7d0 !important; }
    div[class*="stSuccess"] p { color: #14532d !important; }

    div[class*="stError"]   { background-color: #fef2f2 !important; border-color: #fecaca !important; }
    div[class*="stError"] p { color: #7f1d1d !important; }

    div[class*="stWarning"] { background-color: #fffbeb !important; border-color: #fde68a !important; }
    div[class*="stWarning"] p { color: #78350f !important; }

    div[class*="stInfo"]    { background-color: #eff6ff !important; border-color: #bfdbfe !important; }
    div[class*="stInfo"] p  { color: #1e3a8a !important; }

    /* === METRICS === */
    [data-testid="stMetric"] {
        background-color: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1rem 1.25rem;
    }
    [data-testid="stMetricValue"] > div {
        color: #0f172a !important; font-weight: 700 !important; font-size: 1.75rem !important;
    }
    [data-testid="stMetricLabel"] > div {
        color: #64748b !important; font-size: 0.8rem !important;
        text-transform: uppercase; letter-spacing: 0.06em;
    }

    /* === EXPANDER === */
    details > summary { padding: 0.75rem 1rem !important; }
    details > summary p { color: #334155 !important; font-weight: 500 !important; }

    /* === TABS === */
    .stTabs [data-baseweb="tab-list"] { border-bottom: 1px solid #e2e8f0; gap: 2rem; }
    .stTabs [data-baseweb="tab"] { color: #64748b !important; font-weight: 500; padding: 0.75rem 0; }
    .stTabs [data-baseweb="tab"] p { color: inherit !important; }
    .stTabs [aria-selected="true"] { color: #2563eb !important; border-bottom: 2px solid #2563eb !important; background: transparent !important; }

    /* === DATAFRAME === */
    [data-testid="stDataFrame"] { border: 1px solid #e2e8f0; border-radius: 6px; }

    /* === DIVIDER === */
    hr { border-color: #e2e8f0 !important; margin: 2.5rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── Vision Pipeline (inline, no file I/O needed for Streamlit) ────────────────

VISION_PROMPT = """### INPUT ANALYSIS PHASE
1. **Identify Script Style**: First, detect if the text is printed, handwritten, or a mix. Adjust your internal recognition weights for handwritten characters.
2. **Spatial Anchoring**: Use visual lines and margins to define table boundaries.

### EXTRACTION REQUIREMENTS
1. **Hierarchical Extraction**:
   - **Header Data**: Extract titles, dates, and names.
   - **Tabular Structures**: Reconstruct grid-like data.
   - **Currency Detection**: Capture currency symbols exactly.
2. **Handwritten Edge Cases**:
   - Ignore crossed-out text.
   - Use [illegible] for unreadable words.
3. **The "Logical Audit"**:
   - Manually recalculate sums.
   - Report discrepancies as `handwritten_total` and `computed_total`.

### OUTPUT STRUCTURE (JSON)
Constraints:
- Return ONLY valid JSON.
- Do not include comments or markdown formatting (```json).
- "normalized_value" must be a single number or string, NOT a mathematical expression.

{
  "document_summary": { "style": "handwritten/printed", "domain": "auto-detect" },
  "entities": { "label": "value" },
  "tables": [
    {
      "table_description": "string",
      "headers": [],
      "rows": [
        { "column_name": "raw_text", "normalized_value": "numeric_or_string", "currency": "ISO_CODE" }
      ],
      "validation": {
        "math_check": "passed/failed",
        "notes": "string"
      }
    }
  ]
}"""


def get_groq_client():
    """Get Groq client using Streamlit secrets or environment variable."""
    # Try Streamlit secrets first (for Streamlit Cloud deployment)
    try:
        api_key = st.secrets["GROQ_API_KEY"]
    except (KeyError, FileNotFoundError):
        api_key = os.environ.get("GROQ_API_KEY", "")

    if not api_key:
        return None, "GROQ_API_KEY not found. Add it to Streamlit secrets or your .env file."
    return Groq(api_key=api_key), None


def encode_image_bytes(image_bytes: bytes) -> str:
    """Encode image bytes to base64 string."""
    return base64.b64encode(image_bytes).decode("utf-8")


def extract_data_from_image(image_bytes: bytes, filename: str) -> dict | None:
    """Send image to Llama Vision model and return structured JSON data."""
    client, err = get_groq_client()
    if err:
        st.error(f"🔑 API Error: {err}")
        return None

    # Detect MIME type
    ext = filename.lower().rsplit(".", 1)[-1]
    mime = "image/png" if ext == "png" else "image/jpeg"

    b64 = encode_image_bytes(image_bytes)
    image_data_url = f"data:{mime};base64,{b64}"

    try:
        completion = client.chat.completions.create(
            model="meta-llama/llama-4-maverick-17b-128e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": VISION_PROMPT},
                        {"type": "image_url", "image_url": {"url": image_data_url}},
                    ],
                }
            ],
            temperature=0.1,
            response_format={"type": "json_object"},
            stream=False,
        )

        content = completion.choices[0].message.content
        return json.loads(content)

    except json.JSONDecodeError:
        st.error(f"❌ Model returned invalid JSON for **{filename}**.")
        return None
    except Exception as e:
        st.error(f"❌ API Error for **{filename}**: {e}")
        return None



# ── OCR Extraction (EasyOCR + Spatial) ───────────────────────────────────────

import json as _json
import re as _re
from collections import defaultdict

_ocr_reader = None  # lazy-loaded singleton

def _get_ocr_reader():
    """Lazy-load EasyOCR to avoid startup delay."""
    global _ocr_reader
    if _ocr_reader is None:
        import easyocr
        _ocr_reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    return _ocr_reader


def extract_data_from_image_ocr(image_bytes: bytes, filename: str) -> dict | None:
    """
    Run EasyOCR + SpatialTableExtractor inline (no file I/O to disk).
    Returns a dict in the same 'vision-like' schema so build_excel() works
    for both methods.
    """
    try:
        from spatial_table_extractor import SpatialTableExtractor
    except ImportError:
        st.error("SpatialTableExtractor not found. Ensure spatial_table_extractor.py is in the project root.")
        return None

    reader = _get_ocr_reader()
    extractor = SpatialTableExtractor()

    # EasyOCR accepts bytes directly
    import numpy as _np
    from PIL import Image as _PILImage
    import io as _io
    pil_img = _PILImage.open(_io.BytesIO(image_bytes)).convert("RGB")
    img_np = _np.array(pil_img)

    raw_results = reader.readtext(img_np, detail=1, paragraph=False,
                                  min_size=10, text_threshold=0.6,
                                  low_text=0.3, width_ths=0.7, mag_ratio=1.5)

    # Convert to EasyOCR JSON format expected by SpatialTableExtractor
    formatted = [{"bbox": r[0], "text": r[1], "confidence": float(r[2])} for r in raw_results]

    # Write to temp file so SpatialTableExtractor can load it
    with tempfile.NamedTemporaryFile(mode="w", suffix="_easyocr.json", delete=False) as tf:
        import json as _json
        _json.dump(formatted, tf, default=lambda o: o.tolist() if hasattr(o, "tolist") else float(o))
        tmp_path = tf.name

    result = extractor.extract_full_data(tmp_path)

    try:
        import os as _os
        _os.unlink(tmp_path)
    except Exception:
        pass

    if not result:
        return None

    # Convert SpatialTableExtractor output → vision-like schema
    table_data = result.get("table", {})
    headers = table_data.get("headers", [])
    rows_raw = table_data.get("rows", [])

    # Build rows as list of dicts (keyed by header)
    rows_dicts = []
    for row in rows_raw:
        if headers and len(row) == len(headers):
            rows_dicts.append(dict(zip(headers, row)))
        else:
            rows_dicts.append({f"Col {i+1}": v for i, v in enumerate(row)})

    metadata = result.get("metadata", {})
    header_split = result.get("header_split", {})
    footer = result.get("footer_info", [])

    entities = {**metadata}
    for line in (header_split.get("left", []) + header_split.get("right", [])):
        if ":" in line:
            k, _, v = line.partition(":")
            entities[k.strip()] = v.strip()

    footer_rows = [{f"Label": r[0], "Value": r[1]} if len(r) == 2 else {"Text": r[0] if r else ""} for r in footer]

    tables = [{"table_description": f"Table from {filename}", "headers": headers, "rows": rows_dicts,
               "validation": {"math_check": "n/a", "notes": "Extracted via EasyOCR"}}]
    if footer_rows:
        tables.append({"table_description": "Footer / Totals", "headers": list(footer_rows[0].keys()), "rows": footer_rows,
                        "validation": {"math_check": "n/a", "notes": ""}})

    return {
        "document_summary": {"style": "auto-detected", "domain": "auto-detected",
                             "source": "EasyOCR + SpatialTableExtractor"},
        "entities": entities,
        "tables": tables,
    }

# ── Excel Generation ──────────────────────────────────────────────────────────

def build_excel(results: list[dict]) -> bytes:
    """
    Given a list of (sheet_name, vision_data) tuples, build an Excel file
    and return its bytes.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in results:
            doc_summary = data.get("document_summary", {})
            entities = data.get("entities", {})
            tables = data.get("tables", [])

            # Build main table DataFrame
            main_table_df = pd.DataFrame()
            if tables:
                t1 = tables[0]
                rows = t1.get("rows", [])
                if rows:
                    main_table_df = pd.DataFrame(rows)
                else:
                    main_table_df = pd.DataFrame(columns=t1.get("headers", []))

            if main_table_df.empty:
                main_table_df = pd.DataFrame({"Message": ["No tables extracted"]})

            start_table_row = max(len(entities) + len(doc_summary) + 4, 6)
            safe_sheet = sheet_name[:31]  # Excel sheet name limit

            main_table_df.to_excel(writer, sheet_name=safe_sheet, index=False, startrow=start_table_row)
            ws = writer.sheets[safe_sheet]

            # Write metadata
            ws.cell(row=1, column=1, value="DOCUMENT SUMMARY").font = Font(bold=True, size=12)
            r = 2
            for k, v in doc_summary.items():
                ws.cell(row=r, column=1, value=k.title() + ":").font = Font(bold=True)
                ws.cell(row=r, column=2, value=str(v))
                r += 1

            r += 1
            ws.cell(row=r, column=1, value="ENTITIES").font = Font(bold=True, size=12)
            r += 1
            for k, v in entities.items():
                ws.cell(row=r, column=1, value=k.title() + ":").font = Font(bold=True)
                ws.cell(row=r, column=2, value=str(v))
                r += 1

            # Style table header
            header_row = start_table_row + 1
            for col_num, value in enumerate(main_table_df.columns.values):
                cell = ws.cell(row=header_row, column=col_num + 1)
                cell.value = value
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Additional tables
            if len(tables) > 1:
                current_row = header_row + len(main_table_df) + 3
                for t in tables[1:]:
                    ws.cell(row=current_row, column=1, value=t.get("table_description", "Table")).font = Font(bold=True, size=11)
                    current_row += 1
                    headers = t.get("headers", [])
                    t_rows = t.get("rows", [])
                    if not headers and t_rows:
                        headers = list(t_rows[0].keys())
                    for c_idx, h in enumerate(headers):
                        cell = ws.cell(row=current_row, column=c_idx + 1, value=h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="6F819D", end_color="6F819D", fill_type="solid")
                    current_row += 1
                    for row_data in t_rows:
                        for c_idx, h in enumerate(headers):
                            ws.cell(row=current_row, column=c_idx + 1, value=str(row_data.get(h, "")))
                        current_row += 1
                    current_row += 2

            # Auto-size columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    return output.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## Configuration")
    st.markdown("---")

    # Extraction method selector
    st.markdown("### Extraction Method")
    extraction_method = st.radio(
        "Choose extraction engine",
        options=["Groq Vision (Llama)", "EasyOCR (Local)"],
        index=0,
        label_visibility="collapsed",
        help=(
            "**Groq Vision** — Sends images to Llama Vision via Groq API. "
            "Fast, accurate, requires an API key.\n\n"
            "**EasyOCR** — Runs locally, no API key needed. "
            "Downloads models on first run (~250 MB). Best for printed tables."
        ),
    )

    st.markdown("---")

    if extraction_method == "Groq Vision (Llama)":
        # API Key input
        api_key_input = st.text_input(
            "Groq API Key",
            type="password",
            placeholder="gsk_...",
            help="Required if not set in Streamlit secrets or .env file.",
        )
        if api_key_input:
            os.environ["GROQ_API_KEY"] = api_key_input

        st.markdown("**Engine:** `meta-llama/llama-4-maverick-17b-128e-instruct`")
        st.markdown("Powered by **[Groq](https://groq.com)** inference.")
    else:
        st.info(
            "EasyOCR runs entirely on-device. "
            "No API key required. On first use, EasyOCR downloads language models (~250 MB).",
            icon=None,
        )
        st.markdown("**Engine:** EasyOCR + SpatialTableExtractor")

    st.markdown("---")
    st.markdown("### Instructions")
    steps = [
        "1. Choose an extraction method above",
        "2. Upload image files",
        "3. Click Extract Data",
        "4. Review results & download Excel",
    ]
    for s in steps:
        st.markdown(s)

    st.markdown("---")
    st.markdown("### Supported Formats")
    st.markdown("· JPG · JPEG · PNG")


# ── Hero ───────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <h1>ImageToExcel</h1>
    <p>Upload images of invoices, tables, or receipts to automatically extract structured data and generate clean Excel files.</p>
</div>
""", unsafe_allow_html=True)

st.markdown("---")

# ── Upload Section ─────────────────────────────────────────────────────────────

col_upload, col_info = st.columns([2, 1], gap="large")

with col_upload:
    st.markdown("### Upload Documents")
    uploaded_files = st.file_uploader(
        "Drag and drop files here",
        type=["jpg", "jpeg", "png"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

with col_info:
    st.markdown("### Overview")
    m1, m2 = st.columns(2)
    m1.metric("Files Uploaded", len(uploaded_files) if uploaded_files else 0)
    m2.metric("Status", "Ready" if uploaded_files else "Waiting")

st.markdown("---")

# ── Image Preview ──────────────────────────────────────────────────────────────

if uploaded_files:
    st.markdown("### File Preview")
    preview_cols = st.columns(min(len(uploaded_files), 4))
    for i, f in enumerate(uploaded_files):
        with preview_cols[i % 4]:
            st.image(f, caption=f.name, use_container_width=True)

    st.markdown("---")

    # ── Extract Button ─────────────────────────────────────────────────────────
    col_btn, _ = st.columns([1, 2])
    with col_btn:
        run = st.button("Extract Data", type="primary", use_container_width=True)

    if run:
        results = []
        all_raw = {}

        # Progress
        progress_bar = st.progress(0, text="Initializing extraction sequence...")
        status_area = st.empty()

        for idx, uploaded_file in enumerate(uploaded_files):
            pct = int((idx / len(uploaded_files)) * 100)
            progress_bar.progress(pct, text=f"Processing {uploaded_file.name} ({idx+1}/{len(uploaded_files)})")

            with status_area.container():
                engine_label = "Groq Vision" if extraction_method == "Groq Vision (Llama)" else "EasyOCR"
                with st.spinner(f"Analyzing {uploaded_file.name} using {engine_label}..."):
                    image_bytes = uploaded_file.read()
                    t0 = time.time()
                    if extraction_method == "Groq Vision (Llama)":
                        data = extract_data_from_image(image_bytes, uploaded_file.name)
                    else:
                        data = extract_data_from_image_ocr(image_bytes, uploaded_file.name)
                    elapsed = time.time() - t0

            if data:
                sheet_name = uploaded_file.name.rsplit(".", 1)[0][:31]
                results.append((sheet_name, data))
                all_raw[uploaded_file.name] = data
                status_area.success(f"{uploaded_file.name} processed successfully in {elapsed:.1f}s")
            else:
                status_area.error(f"Failed to process {uploaded_file.name}")

        progress_bar.progress(100, text="Extraction complete")

        if results:
            st.markdown("---")
            st.markdown("### Extracted Data")

            tabs = st.tabs([name for name, _ in results])
            for tab, (sheet_name, data) in zip(tabs, results):
                with tab:
                    # Document summary
                    doc_sum = data.get("document_summary", {})
                    entities = data.get("entities", {})
                    tables = data.get("tables", [])

                    c1, c2 = st.columns(2)
                    with c1:
                        if doc_sum:
                            st.markdown("**Document Summary**")
                            st.json(doc_sum)
                    with c2:
                        if entities:
                            st.markdown("**Entities**")
                            st.json(entities)

                    if tables:
                        for t_idx, t in enumerate(tables):
                            desc = t.get("table_description", f"Table {t_idx+1}")
                            st.markdown(f"**{desc}**")
                            rows = t.get("rows", [])
                            if rows:
                                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                            else:
                                st.info("No rows found in this table.")

                            val = t.get("validation", {})
                            if val:
                                check = val.get("math_check", "")
                                notes = val.get("notes", "")
                                badge = "Passed" if "pass" in check.lower() else "Failed"
                                st.caption(f"Validation: {badge} — {notes}")
                    else:
                        st.warning("No tabular data detected in this document.")

                    # Raw JSON expander
                    with st.expander("View Source JSON"):
                        st.json(data)

            # ── Excel Download ─────────────────────────────────────────────────
            st.markdown("---")
            st.markdown("### Export")

            with st.spinner("Generating Excel workbook..."):
                excel_bytes = build_excel(results)

            st.success(f"Workbook generated successfully ({len(results)} sheets)")

            dl_col, _ = st.columns([1, 2])
            with dl_col:
                st.download_button(
                    label="Download Excel File",
                    data=excel_bytes,
                    file_name="Extracted_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        else:
            st.error("Data extraction failed for all uploaded files. Verify your API key configuration and try again.")

else:
    # Empty state
    st.markdown("""
    <div style="text-align:left; padding: 4rem 2rem; color: #64748b; background-color: #ffffff; border: 1px dashed #cbd5e1; border-radius: 8px;">
        <h3 style="color: #334155; font-weight: 500; margin-bottom: 0.5rem;">No documents selected</h3>
        <p style="margin: 0;">Please upload your target documents using the file selector above or drag and drop them into the upload zone.</p>
    </div>
    """, unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:left; color:#94a3b8; font-size:0.875rem;'>"
    "ImageToExcel · Engine: Groq Llama Vision · Interface: Streamlit"
    "</p>",
    unsafe_allow_html=True,
)
