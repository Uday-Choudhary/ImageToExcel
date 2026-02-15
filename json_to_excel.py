"""
JSON to Excel Converter
Converts Vision API JSON output to Excel format.
"""

import os
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

VISION_DATA_DIR = "vision_data"
OUTPUT_DIR = "output"

def save_to_excel(base_dir=".", output_file="Extracted_Data_Vision.xlsx"):
    """
    process extracted header info, tables, and footer info to output/Extracted_Data.xlsx.
    """
    data_dir = os.path.join(base_dir, VISION_DATA_DIR)
    out_dir = os.path.join(base_dir, OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)

    if not os.path.isdir(data_dir):
        print(f"  No {VISION_DATA_DIR}/ folder found. Run Vision pipeline first.")
        return None

    json_files = sorted(
        f for f in os.listdir(data_dir) if f.endswith('_vision.json')
    )
    if not json_files:
        print(f"  No '_vision.json' files in {VISION_DATA_DIR}/. Run vision processor first.")
        return None

    output_path = os.path.join(out_dir, output_file)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        count = 0
        for fname in json_files:
            json_path = os.path.join(data_dir, fname)
            print(f"  Converting {fname} to Excel...")
            
            try:
                with open(json_path, 'r') as f:
                    data = json.load(f)
            except Exception as e:
                print(f"    Error loading {fname}: {e}")
                continue

            if not data:
                continue

            # Parse Data
            doc_summary = data.get("document_summary", {})
            entities = data.get("entities", {})
            tables = data.get("tables", [])
            
            # Sheet name from filename (remove _vision.json)
            sheet_base = fname.replace("_vision.json", "")[:25]
            
            if not tables:
                df = pd.DataFrame({"Message": ["No tables found"]})
                df.to_excel(writer, sheet_name=sheet_base, index=False, startrow=5)
            
            main_table_df = pd.DataFrame()
            if tables:
                t1 = tables[0]
                rows = t1.get("rows", [])
                if rows:
                    main_table_df = pd.DataFrame(rows)
                else:
                    main_table_df = pd.DataFrame(columns=t1.get("headers", []))
            
            start_table_row = max(len(entities) + len(doc_summary) + 4, 6)
            
            main_table_df.to_excel(writer, sheet_name=sheet_base, index=False, startrow=start_table_row)
            ws = writer.sheets[sheet_base]
            
            # --- Write Metadata ---
            ws.cell(row=1, column=1, value="DOCUMENT SUMMARY").font = Font(bold=True)
            r = 2
            for k, v in doc_summary.items():
                ws.cell(row=r, column=1, value=k.title() + ":")
                ws.cell(row=r, column=2, value=str(v))
                r += 1
                
            r += 1
            ws.cell(row=r, column=1, value="ENTITIES").font = Font(bold=True)
            r += 1
            for k, v in entities.items():
                ws.cell(row=r, column=1, value=k.title() + ":")
                ws.cell(row=r, column=2, value=str(v)).font = Font(bold=True)
                r += 1
            
            # --- Style Main Table ---
            header_row = start_table_row + 1
            for col_num, value in enumerate(main_table_df.columns.values):
                cell = ws.cell(row=header_row, column=col_num + 1)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                
            # --- Handle Additional Tables ---
            if len(tables) > 1:
                current_row = header_row + len(main_table_df) + 3
                for t in tables[1:]:
                    ws.cell(row=current_row, column=1, value=t.get("table_description", "Table")).font = Font(bold=True, size=11)
                    current_row += 1
                    
                    headers = t.get("headers", [])
                    t_rows = t.get("rows", [])
                    if not headers and t_rows:
                        headers = list(t_rows[0].keys())
                        
                    for c_idx, h in enumerate(headers):
                        cell = ws.cell(row=current_row, column=c_idx+1, value=h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="6F819D", end_color="6F819D", fill_type="solid")
                    current_row += 1
                    
                    for row_data in t_rows:
                        for c_idx, h in enumerate(headers):
                            val = row_data.get(h, "")
                            ws.cell(row=current_row, column=c_idx+1, value=str(val))
                        current_row += 1
                    current_row += 2

            # Auto-size
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            count += 1

    if count > 0:
        print(f"\n  Saved {count} worksheets to {output_path}")
        return output_path
    else:
        print("\n  No data to save.")
        return None

if __name__ == "__main__":
    save_to_excel()
