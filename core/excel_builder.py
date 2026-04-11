"""
Unified Excel workbook builder for the ImageToExcel pipeline.

Merges the three previous Excel generation implementations
(streamlit_app.py, json_to_excel.py, convert_to_excel.py) into a
single, well-structured module.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.constants import (
    HEADER_FILL_COLOR,
    MAX_COLUMN_WIDTH,
    MAX_SHEET_NAME_LENGTH,
    SECONDARY_FILL_COLOR,
    VALIDATION_FAIL_COLOR,
    VALIDATION_PASS_COLOR,
    MATH_VALIDATION_TOLERANCE,
)

logger = logging.getLogger(__name__)

# Pre-built styles (created once, reused across sheets)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
_SECONDARY_FILL = PatternFill(start_color=SECONDARY_FILL_COLOR, end_color=SECONDARY_FILL_COLOR, fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center")
_PASS_FILL = PatternFill(start_color=VALIDATION_PASS_COLOR, end_color=VALIDATION_PASS_COLOR, fill_type="solid")
_FAIL_FILL = PatternFill(start_color=VALIDATION_FAIL_COLOR, end_color=VALIDATION_FAIL_COLOR, fill_type="solid")


def _safe_sheet_name(name: str) -> str:
    """Truncate and sanitize a string for use as an Excel sheet name.

    Args:
        name: The raw sheet name candidate.

    Returns:
        A string safe for Excel (≤31 chars, no invalid characters).
    """
    # Remove characters invalid in Excel sheet names
    sanitized = re.sub(r'[\\/*?:\[\]]', '_', name)
    return sanitized[:MAX_SHEET_NAME_LENGTH]


def _auto_size_columns(ws) -> None:
    """Auto-size all columns in a worksheet based on cell content.

    Uses `get_column_letter()` to correctly handle columns beyond Z (26+).

    Args:
        ws: An openpyxl worksheet object.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, cell_len)
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, MAX_COLUMN_WIDTH)


def _write_metadata_section(
    ws,
    doc_summary: dict,
    entities: dict,
) -> int:
    """Write document summary and entities to the top of a worksheet.

    Args:
        ws: The openpyxl worksheet.
        doc_summary: Dict of document-level summary fields.
        entities: Dict of extracted entity key-value pairs.

    Returns:
        The next available row number after the metadata section.
    """
    ws.cell(row=1, column=1, value="DOCUMENT SUMMARY").font = Font(bold=True, size=12)
    row = 2
    for key, value in doc_summary.items():
        ws.cell(row=row, column=1, value=f"{key.title()}:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="ENTITIES").font = Font(bold=True, size=12)
    row += 1
    for key, value in entities.items():
        ws.cell(row=row, column=1, value=f"{key.title()}:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    return row


def _style_table_header(ws, header_row: int, columns: list) -> None:
    """Apply styled formatting to a table header row.

    Args:
        ws: The openpyxl worksheet.
        header_row: The 1-indexed row number of the header.
        columns: List of column name strings.
    """
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN


def _write_additional_tables(ws, tables: list[dict], start_row: int) -> None:
    """Write secondary tables (beyond the first) to the worksheet.

    Args:
        ws: The openpyxl worksheet.
        tables: List of table dicts (skipping the first/primary table).
        start_row: Row number to begin writing from.
    """
    current_row = start_row
    for table in tables:
        description = table.get("table_description", "Table")
        ws.cell(row=current_row, column=1, value=description).font = Font(bold=True, size=11)
        current_row += 1

        headers = table.get("headers", [])
        rows = table.get("rows", [])
        if not headers and rows:
            headers = list(rows[0].keys())

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _SECONDARY_FILL
        current_row += 1

        for row_data in rows:
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col_idx, value=str(row_data.get(header, "")))
            current_row += 1
        current_row += 2


def _add_math_validation(
    ws,
    headers: list[str],
    rows: list,
    header_row: int,
) -> None:
    """Add a Validation column with Qty × Price vs Total checks.

    Args:
        ws: The openpyxl worksheet.
        headers: List of column header strings.
        rows: List of row data (list of dicts or list of lists).
        header_row: The 1-indexed row of the table header.
    """
    col_map = {name.lower(): idx for idx, name in enumerate(headers)}

    qty_idx = price_idx = total_idx = -1
    for name, idx in col_map.items():
        if any(kw in name for kw in ("qty", "quantity", "units")):
            qty_idx = idx
        if any(kw in name for kw in ("price", "rate", "unit")):
            price_idx = idx
        if any(kw in name for kw in ("total", "amount", "net")):
            total_idx = idx

    if qty_idx == -1 or price_idx == -1 or total_idx == -1:
        return

    val_col = len(headers) + 1
    ws.cell(row=header_row, column=val_col, value="Validation").font = Font(bold=True)

    for row_i, row_data in enumerate(rows):
        current_row = header_row + 1 + row_i
        try:
            def _extract_number(val: str) -> float:
                cleaned = re.sub(r'[^\d.]', '', str(val))
                return float(cleaned) if cleaned else 0.0

            q = _extract_number(row_data[qty_idx] if isinstance(row_data, list) else row_data.get(headers[qty_idx], ""))
            p = _extract_number(row_data[price_idx] if isinstance(row_data, list) else row_data.get(headers[price_idx], ""))
            t = _extract_number(row_data[total_idx] if isinstance(row_data, list) else row_data.get(headers[total_idx], ""))

            calc = q * p
            diff = abs(calc - t)

            status_cell = ws.cell(row=current_row, column=val_col)
            if diff < MATH_VALIDATION_TOLERANCE and calc > 0:
                status_cell.value = "OK"
                status_cell.fill = _PASS_FILL
            elif calc > 0:
                status_cell.value = f"Mismatch (Calc: {calc:.2f})"
                status_cell.fill = _FAIL_FILL
                ws.cell(row=current_row, column=total_idx + 1).fill = _FAIL_FILL
        except (ValueError, TypeError, IndexError):
            pass


def build_excel_from_vision(results: list[tuple[str, dict]]) -> bytes:
    """Build an Excel workbook from Vision API extraction results.

    Each (sheet_name, data) tuple becomes a sheet in the workbook with:
    - Document summary and entities at the top
    - Primary table with styled headers
    - Additional tables below
    - Math validation column where applicable
    - Auto-sized columns

    Args:
        results: List of (sheet_name, vision_data_dict) tuples.

    Returns:
        The Excel file contents as bytes.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, data in results:
            doc_summary = data.get("document_summary", {})
            entities = data.get("entities", {})
            tables = data.get("tables", [])

            # Build primary table DataFrame
            main_df = pd.DataFrame()
            if tables:
                first_table = tables[0]
                rows = first_table.get("rows", [])
                main_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=first_table.get("headers", []))

            if main_df.empty:
                main_df = pd.DataFrame({"Message": ["No tables extracted"]})

            start_row = max(len(entities) + len(doc_summary) + 4, 6)
            safe_name = _safe_sheet_name(sheet_name)

            main_df.to_excel(writer, sheet_name=safe_name, index=False, startrow=start_row)
            ws = writer.sheets[safe_name]

            # Write metadata
            _write_metadata_section(ws, doc_summary, entities)

            # Style table header
            header_row = start_row + 1
            _style_table_header(ws, header_row, list(main_df.columns))

            # Additional tables
            if len(tables) > 1:
                extra_start = header_row + len(main_df) + 3
                _write_additional_tables(ws, tables[1:], extra_start)

            # Auto-size
            _auto_size_columns(ws)

            logger.info("Sheet '%s' created with %d rows", safe_name, len(main_df))

    return output.getvalue()


def build_excel_from_ocr(
    sheets_data: list[dict],
) -> Optional[bytes]:
    """Build an Excel workbook from OCR spatial extraction results.

    Each dict in `sheets_data` should contain:
    - sheet_name: str
    - header_split: {"left": [...], "right": [...]}
    - metadata: dict
    - table: {"headers": [...], "rows": [...]}
    - footer_info: list

    Args:
        sheets_data: List of extraction result dicts.

    Returns:
        The Excel file contents as bytes, or None if no data.
    """
    if not sheets_data:
        return None

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        count = 0
        for entry in sheets_data:
            sheet_name = entry.get("sheet_name", f"Sheet{count + 1}")
            header_split = entry.get("header_split", {"left": [], "right": []})
            metadata = entry.get("metadata", {})
            table = entry.get("table", {})
            footer_info = entry.get("footer_info", [])

            headers = table.get("headers", [])
            rows = table.get("rows", [])

            if not rows:
                continue

            safe_name = _safe_sheet_name(sheet_name)

            # Create DataFrame
            df = pd.DataFrame(rows, columns=headers)
            header_height = max(len(header_split.get("left", [])), len(header_split.get("right", [])))
            start_row = max(header_height + 4, 6)

            df.to_excel(writer, sheet_name=safe_name, index=False, startrow=start_row)
            ws = writer.sheets[safe_name]

            # Write left header
            for idx, text in enumerate(header_split.get("left", [])):
                cell = ws.cell(row=idx + 1, column=1, value=text)
                cell.font = Font(bold=(idx == 0))

            # Write right header
            right_col = max(len(headers), 6)
            for idx, text in enumerate(header_split.get("right", [])):
                cell = ws.cell(row=idx + 1, column=right_col, value=text)
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(bold=(idx == 0))

            # Write metadata
            if metadata:
                meta_row, meta_col = 1, 3
                ws.cell(row=meta_row, column=meta_col, value="METADATA EXTRACTED").font = Font(bold=True, underline="single")
                for key, value in metadata.items():
                    meta_row += 1
                    ws.cell(row=meta_row, column=meta_col, value=f"{key.replace('_', ' ').title()}:")
                    ws.cell(row=meta_row, column=meta_col + 1, value=value).font = Font(bold=True)

            # Style table header
            header_row_idx = start_row + 1
            _style_table_header(ws, header_row_idx, headers)

            # Math validation
            _add_math_validation(ws, headers, rows, header_row_idx)

            # Write footer info
            footer_start = header_row_idx + len(rows) + 2
            for idx, info_row in enumerate(footer_info):
                target_col_val = len(headers)
                target_col_label = max(1, target_col_val - 1)
                row_num = footer_start + idx

                if len(info_row) >= 2:
                    ws.cell(row=row_num, column=target_col_label, value=info_row[0]).alignment = Alignment(horizontal="right")
                    ws.cell(row=row_num, column=target_col_val, value=info_row[1]).alignment = Alignment(horizontal="right")
                    if "total" in info_row[0].lower() or "due" in info_row[0].lower():
                        ws.cell(row=row_num, column=target_col_label).font = Font(bold=True)
                        ws.cell(row=row_num, column=target_col_val).font = Font(bold=True)
                else:
                    ws.cell(row=row_num, column=1, value=info_row[0])

            # Auto-size
            _auto_size_columns(ws)

            logger.info(
                "Sheet '%s' | %d left / %d right headers | %d rows",
                safe_name,
                len(header_split.get("left", [])),
                len(header_split.get("right", [])),
                len(rows),
            )
            count += 1

    if count == 0:
        return None

    return output.getvalue()
