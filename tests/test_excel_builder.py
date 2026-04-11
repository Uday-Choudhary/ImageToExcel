"""Tests for the unified Excel builder module."""

from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.excel_builder import (
    _auto_size_columns,
    _safe_sheet_name,
    build_excel_from_vision,
)


class TestSafeSheetName:
    """Tests for Excel sheet name sanitization."""

    def test_normal_name(self) -> None:
        assert _safe_sheet_name("Invoice") == "Invoice"

    def test_truncates_long_names(self) -> None:
        long_name = "A" * 50
        assert len(_safe_sheet_name(long_name)) == 31

    def test_removes_invalid_characters(self) -> None:
        assert _safe_sheet_name("Sheet[1]") == "Sheet_1_"
        assert _safe_sheet_name("A/B\\C") == "A_B_C"
        assert _safe_sheet_name("Test*Name?") == "Test_Name_"

    def test_empty_string(self) -> None:
        assert _safe_sheet_name("") == ""


class TestBuildExcelFromVision:
    """Tests for Vision-mode Excel generation."""

    def test_creates_valid_workbook(self, sample_vision_results) -> None:
        excel_bytes = build_excel_from_vision(sample_vision_results)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "TestInvoice"

    def test_writes_document_summary(self, sample_vision_results) -> None:
        excel_bytes = build_excel_from_vision(sample_vision_results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "DOCUMENT SUMMARY"

    def test_writes_entities(self, sample_vision_results) -> None:
        excel_bytes = build_excel_from_vision(sample_vision_results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Find ENTITIES header
        found = False
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1):
            for cell in row:
                if cell.value == "ENTITIES":
                    found = True
                    break
        assert found

    def test_writes_table_data(self, sample_vision_results) -> None:
        excel_bytes = build_excel_from_vision(sample_vision_results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Find "Description" header in the table
        found = False
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
            for cell in row:
                if cell.value == "Description":
                    found = True
                    break
        assert found

    def test_empty_tables_shows_message(self) -> None:
        results = [("Empty", {"document_summary": {}, "entities": {}, "tables": []})]
        excel_bytes = build_excel_from_vision(results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Should contain "No tables extracted" message
        found = False
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=5):
            for cell in row:
                if cell.value == "No tables extracted":
                    found = True
                    break
        assert found

    def test_multiple_sheets(self, sample_vision_result) -> None:
        results = [
            ("Sheet1", sample_vision_result),
            ("Sheet2", sample_vision_result),
        ]
        excel_bytes = build_excel_from_vision(results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 2

    def test_styled_header_row(self, sample_vision_results) -> None:
        """Table headers should have white font (styled)."""
        excel_bytes = build_excel_from_vision(sample_vision_results)
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Find the row containing "Description" header
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
            for cell in row:
                if cell.value == "Description" and cell.font.color:
                    assert cell.font.bold is True
                    break
